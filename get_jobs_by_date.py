"""Return jobs expected to be in production on a given date.

Each sheet in the workbook represents a batch of jobs that started on a
particular day.  A job is considered "active" on *target_date* when:

    start_date <= target_date <= end_date

where start_date / end_date are derived from the "Start Time" / "End Time"
columns (column I and K, i.e. zero-based indices 8 and 10).
"""

from __future__ import annotations

from datetime import date, datetime
from typing import Any, TypeAlias

import openpyxl

_DateLike: TypeAlias = str | datetime | date | None

_DATE_FORMATS = ("%d-%m-%Y %H:%M", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d")


def _parse_str_to_date(value: str) -> date | None:
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except ValueError:
            pass
    return None


def _to_date(value: _DateLike) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return _parse_str_to_date(value)
    return None


def _find_header_row(rows: list[Any]) -> int | None:
    for idx, row in enumerate(rows):
        if row and row[0] == "Order No.":
            return idx
    return None


def _order_no_if_active(row: Any, target_date: date) -> str | None:
    if not row or row[0] is None:
        return None
    start = _to_date(row[8])
    end = _to_date(row[10])
    if start is None or end is None:
        return None
    if start <= target_date <= end:
        return str(row[0]).strip()
    return None


def _jobs_in_sheet(ws: Any, target_date: date) -> list[str]:
    rows: list[Any] = list(ws.iter_rows(values_only=True))
    header_idx = _find_header_row(rows)
    if header_idx is None:
        return []
    jobs: list[str] = []
    for row in rows[header_idx + 1 :]:
        order_no = _order_no_if_active(row, target_date)
        if order_no is not None:
            jobs.append(order_no)
    return jobs


def get_jobs_by_date(filepath: str, target_date: date) -> list[str]:
    """Return a list of Order Numbers whose operation spans *target_date*.

    Parameters
    ----------
    filepath:
        Path to the "Operations by day" Excel workbook.
    target_date:
        The date to check against each job's Start Time / End Time.

    Returns
    -------
    list[str]
        De-duplicated list of Order Numbers active on *target_date*, in the
        order they were encountered while scanning the sheets.
    """
    wb: Any = openpyxl.load_workbook(filepath, data_only=True)
    seen: set[str] = set()
    jobs: list[str] = []
    for sheet_name in wb.sheetnames:
        for order_no in _jobs_in_sheet(wb[sheet_name], target_date):
            if order_no not in seen:
                seen.add(order_no)
                jobs.append(order_no)
    return jobs
