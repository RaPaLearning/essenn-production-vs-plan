import datetime
from typing import Any, cast

import openpyxl
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule  # type: ignore[reportUnknownVariableType]
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from get_active_jobs import get_active_jobs
from setup_cycle_times import (
    CycleTimeLookup,
    MachineTypeLookup,
    load_cycle_times,
    load_machine_types,
)

_SHIFT_TIMINGS: dict[str, tuple[str, str]] = {
    "A": ("Turning: 6:00 AM \u2013 2:00 PM  |  Milling/Citizen: 6:00 AM \u2013 2:30 PM", ""),
    "B": ("Turning: 2:00 PM \u2013 10:00 PM  |  Milling/Citizen: 2:30 PM \u2013 11:00 PM", ""),
    "C": ("Turning: 10:00 PM \u2013 6:00 AM", ""),
}

_HEADERS: dict[str, str] = {
    "A6": "S.No.",
    "B6": "MACHINE",
    "C6": "JOB ORDER No",
    "D6": "TOTAL QTY",
    "E6": "PART NO",
    "F6": "PART NAME",
    "G6": "OPERATION",
    "H6": "PLAN QTY",
    "I6": "ACTUAL QTY",
    "I7": "OK QTY",
    "J7": "Rej QTY",
    "K6": "%",
    "L6": "REMARKS",
    "M6": "Operator Name",
    "N6": "Sign",
}
_NUM_COLS: int = max(column_index_from_string(ref.rstrip("0123456789")) for ref in _HEADERS)
_LAST_COL: str = get_column_letter(_NUM_COLS)


def _make_border(
    thin: Side,
    thick: Side,
    col: int,
    is_top_thick: bool,
    is_bottom_thick: bool,
) -> Border:
    """Create a cell border with thick edges on row/column boundaries."""
    return Border(
        left=thick if col == 1 else thin,
        right=thick if col == _NUM_COLS else thin,
        top=thick if is_top_thick else thin,
        bottom=thick if is_bottom_thick else thin,
    )


def _write_header(
    ws: openpyxl.worksheet.worksheet.Worksheet,  # type: ignore[reportUnknownMemberType]
    label: str,
    target_date: datetime.date,
    thin: Side,
    thick: Side,
    meta_font: Font,
) -> None:
    """Write the DATE and SHIFT header rows."""
    formatted_date = target_date.strftime("%d-%m-%Y")
    bold_font = Font(name="Arial", size=10, bold=True)

    ws.merge_cells("A1:B1")  # type: ignore[reportUnknownMemberType]
    ws["A1"] = "DATE:"  # type: ignore[reportUnknownMemberType]
    ws["A1"].font = bold_font  # type: ignore[reportUnknownMemberType]
    ws.merge_cells(f"C1:{_LAST_COL}1")  # type: ignore[reportUnknownMemberType]
    ws["C1"] = formatted_date  # type: ignore[reportUnknownMemberType]
    ws["C1"].font = meta_font  # type: ignore[reportUnknownMemberType]
    ws["C1"].alignment = Alignment(horizontal="left", vertical="center", shrink_to_fit=True)  # type: ignore[reportUnknownMemberType]

    ws.merge_cells("A2:B2")  # type: ignore[reportUnknownMemberType]
    ws["A2"] = "SHIFT:"  # type: ignore[reportUnknownMemberType]
    ws["A2"].font = bold_font  # type: ignore[reportUnknownMemberType]
    ws.merge_cells(f"C2:{_LAST_COL}2")  # type: ignore[reportUnknownMemberType]
    ws["C2"] = f"Shift {label}"  # type: ignore[reportUnknownMemberType]
    ws["C2"].font = meta_font  # type: ignore[reportUnknownMemberType]
    ws["C2"].alignment = Alignment(horizontal="left", vertical="center", shrink_to_fit=True)  # type: ignore[reportUnknownMemberType]

    for r in [1, 2]:
        for c in range(1, _NUM_COLS + 1):
            ws.cell(row=r, column=c).border = _make_border(  # type: ignore[reportUnknownMemberType]
                thin,
                thick,
                c,
                r == 1,
                r == 2,
            )

    # Shift timing note in rows 3-4
    timing_text = _SHIFT_TIMINGS.get(label, ("", ""))[0]
    if timing_text:
        note_font = Font(name="Arial", size=8, italic=True, color="555555")
        ws.merge_cells(f"A3:{_LAST_COL}3")  # type: ignore[reportUnknownMemberType]
        ws["A3"] = timing_text  # type: ignore[reportUnknownMemberType]
        ws["A3"].font = note_font  # type: ignore[reportUnknownMemberType]
        ws["A3"].alignment = Alignment(horizontal="left", vertical="center", shrink_to_fit=True)  # type: ignore[reportUnknownMemberType]


def _style_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,  # type: ignore[reportUnknownMemberType]
    r: int,
    font: Font,
    thin: Side,
    thick: Side,
    is_top_thick: bool,
    is_bottom_thick: bool,
    fill: PatternFill | None = None,
    align: Alignment | None = None,
) -> None:
    """Apply uniform styling to all columns of a single row."""
    for c in range(1, _NUM_COLS + 1):
        cell = ws.cell(row=r, column=c)  # type: ignore[reportUnknownMemberType]
        cell.font = font  # type: ignore[reportUnknownMemberType]
        cell.border = _make_border(  # type: ignore[reportUnknownMemberType]
            thin,
            thick,
            c,
            is_top_thick,
            is_bottom_thick,
        )
        if fill is not None:
            cell.fill = fill  # type: ignore[reportUnknownMemberType]
        if align is not None:
            cell.alignment = align  # type: ignore[reportUnknownMemberType]


def _write_table_headers(
    ws: openpyxl.worksheet.worksheet.Worksheet,  # type: ignore[reportUnknownMemberType]
    header_font: Font,
    fill_grey: PatternFill,
    thin: Side,
    thick: Side,
) -> None:
    """Write the column headers in rows 6-7."""
    vertical_merges = [
        "A6:A7",
        "B6:B7",
        "C6:C7",
        "D6:D7",
        "E6:E7",
        "F6:F7",
        "G6:G7",
        "H6:H7",
        "K6:K7",
        "L6:L7",
        "M6:M7",
        "N6:N7",
    ]
    for m in vertical_merges:
        ws.merge_cells(m)  # type: ignore[reportUnknownMemberType]
    ws.merge_cells("I6:J6")  # type: ignore[reportUnknownMemberType]

    for cell_ref, value in _HEADERS.items():
        ws[cell_ref] = value  # type: ignore[reportUnknownMemberType]

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
        shrink_to_fit=True,
    )
    for r in [6, 7]:
        ws.row_dimensions[r].height = 25  # type: ignore[reportUnknownMemberType]
        _style_row(
            ws,  # type: ignore[reportUnknownArgumentType]
            r,
            header_font,
            thin,
            thick,
            r == 6,
            r == 7,
            fill_grey,
            center,
        )


def _write_single_data_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,  # type: ignore[reportUnknownMemberType]
    r_num: int,
    idx: int,
    row: dict[str, Any],
    data_font: Font,
    thin: Side,
    thick: Side,
    is_last: bool,
) -> None:
    """Write one job data row and style its cells."""
    data_keys = [
        "Machine",
        "Job Order No",
        "Total Qty",
        "Part No",
        "Part Name",
        "Operation",
        "Plan Qty",
    ]
    is_anomaly = "Anomaly" in row

    ws.row_dimensions[r_num].height = 24  # type: ignore[reportUnknownMemberType]
    ws.cell(row=r_num, column=1, value=idx + 1)  # type: ignore[reportUnknownMemberType]
    for col_i, key in enumerate(data_keys, start=2):
        ws.cell(row=r_num, column=col_i, value=row.get(key, ""))  # type: ignore[reportUnknownMemberType]

    # Efficiency % formula in column K (11): auto-calculates when OK QTY is entered
    eff_formula = f'=IF(H{r_num}>0,ROUND(I{r_num}/H{r_num}*100,0),"")'
    ws.cell(row=r_num, column=11, value=eff_formula)  # type: ignore[reportUnknownMemberType]

    for c in [9, 10, 12, 13, 14]:
        ws.cell(row=r_num, column=c, value="")  # type: ignore[reportUnknownMemberType]

    if is_anomaly:
        fill_pink = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        plan_cell = ws.cell(row=r_num, column=8)  # type: ignore[reportUnknownMemberType]
        plan_cell.fill = fill_pink  # type: ignore[reportUnknownMemberType]
        plan_cell.comment = Comment(row["Anomaly"], "System")  # type: ignore[reportUnknownMemberType]

    _style_data_cells(ws, r_num, data_font, thin, thick, is_last)  # type: ignore[reportUnknownArgumentType]


def _style_data_cells(
    ws: openpyxl.worksheet.worksheet.Worksheet,  # type: ignore[reportUnknownMemberType]
    r_num: int,
    data_font: Font,
    thin: Side,
    thick: Side,
    is_last: bool,
) -> None:
    """Apply font, border, and alignment to a data row."""
    center_cols = {1, 4, 8, 9, 10, 11}
    for c in range(1, _NUM_COLS + 1):
        cell = ws.cell(row=r_num, column=c)  # type: ignore[reportUnknownMemberType]
        cell.font = data_font  # type: ignore[reportUnknownMemberType]
        cell.border = _make_border(thin, thick, c, False, is_last)  # type: ignore[reportUnknownMemberType]
        h = "center" if c in center_cols else "left"
        cell.alignment = Alignment(horizontal=h, vertical="center", shrink_to_fit=True)  # type: ignore[reportUnknownMemberType]


def _write_empty_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,  # type: ignore[reportUnknownMemberType]
    start_row: int,
    data_font: Font,
    thin: Side,
    thick: Side,
) -> None:
    """Write a placeholder row when no jobs are active."""
    ws.row_dimensions[start_row].height = 24  # type: ignore[reportUnknownMemberType]
    ws.cell(row=start_row, column=1, value="-")  # type: ignore[reportUnknownMemberType]
    ws.cell(  # type: ignore[reportUnknownMemberType]
        row=start_row,
        column=2,
        value="No active jobs scheduled for this shift",
    )
    ws.merge_cells(  # type: ignore[reportUnknownMemberType]
        start_row=start_row,
        start_column=2,
        end_row=start_row,
        end_column=_NUM_COLS,
    )
    _style_row(
        ws,  # type: ignore[reportUnknownArgumentType]
        start_row,
        data_font,
        thin,
        thick,
        False,
        True,
    )


def _write_data_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,  # type: ignore[reportUnknownMemberType]
    rows: list[dict[str, Any]],
    data_font: Font,
    thin: Side,
    thick: Side,
) -> int:
    """Write job data rows starting at row 8. Returns last grid row."""
    start_row = 8
    for idx, row in enumerate(rows):
        _write_single_data_row(
            ws,  # type: ignore[reportUnknownArgumentType]
            start_row + idx,
            idx,
            row,
            data_font,
            thin,
            thick,
            is_last=(idx == len(rows) - 1),
        )

    if rows:
        return start_row + len(rows) - 1

    _write_empty_row(ws, start_row, data_font, thin, thick)  # type: ignore[reportUnknownArgumentType]
    return start_row


def _write_footer(
    ws: openpyxl.worksheet.worksheet.Worksheet,  # type: ignore[reportUnknownMemberType]
    last_grid_row: int,
) -> None:
    """Write signature footer and set column widths."""
    foot_row = last_grid_row + 3
    sign_font = Font(name="Arial", size=10, bold=True)
    ws.cell(  # type: ignore[reportUnknownMemberType]
        row=foot_row,
        column=2,
        value="Sign Of Incharge :",
    ).font = sign_font
    ws.cell(  # type: ignore[reportUnknownMemberType]
        row=foot_row,
        column=10,
        value="Sign Of Manager :",
    ).font = sign_font

    col_widths = {
        1: 6,
        2: 18,
        3: 16,
        4: 12,
        5: 14,
        6: 45,
        7: 18,
        8: 12,
        9: 10,
        10: 10,
        11: 8,
        12: 16,
        13: 18,
        14: 10,
    }
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w  # type: ignore[reportUnknownMemberType]


def _fill_shift_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,  # type: ignore[reportUnknownMemberType]
    label: str,
    target_date: datetime.date,
    rows: list[dict[str, Any]],
    anomalies: list[dict[str, Any]],
) -> None:
    """Style and fill the given sheet with the shift template layout."""
    ws.views.sheetView[0].showGridLines = True  # type: ignore[reportUnknownMemberType]

    # Configure print settings: landscape and fit all columns to one page wide
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # type: ignore[reportUnknownMemberType]
    ws.sheet_properties.pageSetUpPr.fitToPage = True  # type: ignore[reportUnknownMemberType]
    ws.page_setup.fitToWidth = 1  # type: ignore[reportUnknownMemberType]
    ws.page_setup.fitToHeight = 1  # type: ignore[reportUnknownMemberType]

    header_font = Font(name="Arial", size=10, bold=True)
    meta_font = Font(name="Arial", size=9)
    data_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="000000")
    thick = Side(style="medium", color="000000")
    fill_grey = PatternFill(
        start_color="F2F2F2",
        end_color="F2F2F2",
        fill_type="solid",
    )

    _write_header(ws, label, target_date, thin, thick, meta_font)  # type: ignore[reportUnknownArgumentType]
    _write_table_headers(ws, header_font, fill_grey, thin, thick)  # type: ignore[reportUnknownArgumentType]

    # Combine normal rows and anomalies into one grid, sorted by machine
    all_rows = rows + anomalies
    all_rows.sort(key=lambda r: str(r.get("Machine", "")))

    last_row = _write_data_rows(ws, all_rows, data_font, thin, thick)  # type: ignore[reportUnknownArgumentType]

    # Conditional formatting: yellow fill when Efficiency % < 85
    if all_rows:
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        eff_rule = FormulaRule(  # type: ignore[reportUnknownVariableType]
            formula=["AND(ISNUMBER(K8),K8<85)"],
            stopIfTrue=True,
            fill=yellow_fill,
        )
        ws.conditional_formatting.add(f"K8:K{last_row}", eff_rule)  # type: ignore[reportUnknownMemberType, reportUnknownArgumentType]

    _write_footer(ws, last_row)  # type: ignore[reportUnknownArgumentType]


def export_active_jobs(
    input_path: str,
    date_str: str,
    output_path: str,
    masterlist_path: str | None = None,
) -> None:
    """Export active jobs and pre-fill shift-wise template sheets to the output Excel file."""
    target_date: datetime.date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()

    cycle_lookup: CycleTimeLookup | None = None
    machine_type_lookup: MachineTypeLookup | None = None
    if masterlist_path is not None:
        cycle_lookup = load_cycle_times(masterlist_path)
        machine_type_lookup = load_machine_types(masterlist_path)

    rows, anomalies = get_active_jobs(
        input_path,
        date_str,
        cycle_lookup,
        machine_type_lookup,
    )

    # Save workbook with exactly the three shift sheets
    wb = openpyxl.Workbook()

    # Configure Shift A on the default first sheet
    ws_shift_a: openpyxl.worksheet.worksheet.Worksheet = (  # type: ignore[reportUnknownMemberType]
        cast(openpyxl.worksheet.worksheet.Worksheet, wb.active)  # type: ignore[reportUnknownMemberType]
    )
    ws_shift_a.title = "Shift A"  # type: ignore[reportUnknownMemberType]
    rows_a = [r for r in rows if r.get("Shift") == "Shift A"]
    anom_a = [a for a in anomalies if a.get("Shift") == "Shift A"]
    _fill_shift_sheet(ws_shift_a, "A", target_date, rows_a, anom_a)  # type: ignore[reportUnknownArgumentType]

    # Create Shift B and C sheets
    ws_shift_b = wb.create_sheet(title="Shift B")  # type: ignore[reportUnknownMemberType]
    rows_b = [r for r in rows if r.get("Shift") == "Shift B"]
    anom_b = [a for a in anomalies if a.get("Shift") == "Shift B"]
    _fill_shift_sheet(ws_shift_b, "B", target_date, rows_b, anom_b)

    ws_shift_c = wb.create_sheet(title="Shift C")  # type: ignore[reportUnknownMemberType]
    rows_c = [r for r in rows if r.get("Shift") == "Shift C"]
    anom_c = [a for a in anomalies if a.get("Shift") == "Shift C"]
    _fill_shift_sheet(ws_shift_c, "C", target_date, rows_c, anom_c)

    wb.save(output_path)  # type: ignore[reportUnknownMemberType]
