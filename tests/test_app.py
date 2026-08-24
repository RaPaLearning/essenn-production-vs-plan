import io
import logging
import os
import re
import unittest
from datetime import date

import openpyxl
import pandas as pd
from streamlit.testing.v1 import AppTest

# Suppress all Streamlit stderr noise (logging + raw pyarrow tracebacks).
logging.getLogger("streamlit").setLevel(logging.ERROR)


class _SilenceStderr(unittest.TestCase):
    """Shared base that redirects stderr to devnull during tests."""

    def setUp(self) -> None:
        """Redirect the OS-level stderr fd to devnull to silence pyarrow tracebacks."""
        self._orig_stderr_fd = os.dup(2)
        self._devnull_fd = os.open(os.devnull, os.O_WRONLY)
        os.dup2(self._devnull_fd, 2)

    def tearDown(self) -> None:
        """Restore the OS-level stderr fd after each test."""
        os.dup2(self._orig_stderr_fd, 2)
        os.close(self._orig_stderr_fd)
        os.close(self._devnull_fd)


class TestApp(_SilenceStderr):
    def test_app_loads_and_processes_file(self):
        """Test the Streamlit app logic from loading to file download using AppTest naturally."""
        # Initialize the app test pointing to our main app script
        at = AppTest.from_file("app.py", default_timeout=30)
        at.run()

        # Verify that the app loaded correctly without any immediate crashes
        self.assertFalse(at.exception, f"App failed to load: {at.exception}")

        # Verify main UI elements are present
        self.assertEqual(at.title[0].value, "📊 Operations Extractor")

        # Verify baseline UI state (no download button should be shown initially)
        has_unknown_elements = any(
            getattr(e, "type", type(e).__name__) == "UnknownElement" for e in at.main
        )
        self.assertFalse(
            has_unknown_elements,
            "Download button element should not exist before an upload is completed",
        )
        self.assertEqual(len(at.subheader), 0, "Preview subheader should not exist initially")
        self.assertEqual(len(at.dataframe), 0, "Dataframe preview should not exist initially")

        # Find sample data
        sample_path = "data/sample/OperationsByDay.xlsx"
        self.assertTrue(os.path.exists(sample_path), f"Sample file not found at {sample_path}")

        with open(sample_path, "rb") as f:
            file_bytes = f.read()

        # Set the target date and trigger a stream re-run
        target_date = date(2026, 3, 16)
        at.date_input[0].set_value(target_date).run()
        self.assertFalse(at.exception, "App crashed while setting date")

        # Simulate uploading the file via the file uploader component
        # Streamlit AppTest expects a list of tuples containing (filename, content_bytes, mime_type)
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        at.file_uploader[0].set_value([("OperationsByDay.xlsx", file_bytes, mime)]).run()

        # Verify that processing the mock file didn't crash the Streamlit app script
        self.assertFalse(at.exception, f"App crashed during file processing: {at.exception}")

        # The main validation: check if the download button (UnknownElement subtype)
        # was appended to the bottom and it only renders if the file processing runs completely
        # without hitting earlier exceptions and correctly saves bytes to session_state.
        has_download_button = any(
            getattr(e, "type", type(e).__name__) == "download_button" for e in at.main
        )
        self.assertTrue(
            has_download_button,
            "Download button element was not rendered into the tree (processing failed)",
        )

        # Verify the preview subheader and dataframe element are rendered successfully
        self.assertGreater(len(at.subheader), 0, "Preview subheader should be rendered")
        self.assertEqual(at.subheader[0].value, "Preview Summary")
        self.assertGreater(len(at.dataframe), 0, "Dataframe preview should be rendered")

    def test_app_handles_invalid_file(self):
        """Test that uploading a completely invalid file gets caught naturally.

        This tests the except block to ensure there are no crashes in AppTest.
        """
        at = AppTest.from_file("app.py", default_timeout=30)
        at.run()

        # Ensure we are not on a Sunday by setting a known Monday
        at.date_input[0].set_value(date(2026, 3, 16)).run()

        # Uploading raw text bytes mapped to an excel mime to trigger pandas crash artificially
        bad_bytes = b"This is definitely not an excel file."
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        at.file_uploader[0].set_value([("bad_data.xlsx", bad_bytes, mime)]).run()

        # Verify the application itself did NOT crash natively
        self.assertFalse(
            at.exception,
            f"App completely crashed instead of gracefully handling the exception: {at.exception}",
        )

        # It should hit the st.error block
        self.assertGreater(
            len(at.error), 0, "Expected an st.error block to render when processing fails"
        )
        self.assertIn(
            "Error processing file",
            at.error[0].value,
            "The error message did not match expectations",
        )

        # The download button must NOT exist.
        has_download_button = any(
            getattr(e, "type", type(e).__name__) == "download_button" for e in at.main
        )
        self.assertFalse(
            has_download_button, "A download button was improperly rendered after a failure"
        )

        # The preview components must NOT exist
        self.assertEqual(
            len(at.subheader), 0, "Preview subheader should not be rendered after a failure"
        )
        self.assertEqual(
            len(at.dataframe), 0, "Dataframe preview should not be rendered after a failure"
        )

    def test_sidebar_version_display(self) -> None:
        """Test that a version number matching x.y is displayed in the sidebar."""
        at = AppTest.from_file("app.py", default_timeout=30)
        at.run()
        self.assertFalse(at.exception, f"App failed to load: {at.exception}")
        captions = [c.value for c in at.sidebar.caption]
        self.assertTrue(
            any(re.search(r"\d+\.\d+", cap) for cap in captions),
            f"Version matching x.y not found in sidebar captions: {captions}",
        )

    def test_app_blocks_sunday(self):
        """Selecting a Sunday should show an error and stop the app."""
        at = AppTest.from_file("app.py", default_timeout=30)
        at.run()

        # Set date to a Sunday (July 5, 2026 is a Sunday)
        at.date_input[0].set_value(date(2026, 7, 5)).run()
        self.assertFalse(at.exception, "App crashed when selecting a Sunday")

        # Verify the Sunday error message is shown
        self.assertGreater(len(at.error), 0, "Expected an error when Sunday is selected")
        self.assertIn(
            "Sundays are not working days",
            at.error[0].value,
        )


class TestIntegratorPage(_SilenceStderr):
    """Tests for the Actuals Integrator page."""

    def test_integrator_page_loads(self) -> None:
        """The integrator page should load without errors."""
        at = AppTest.from_file(
            "actuals_integrator_page.py",
            default_timeout=30,
        )
        at.run()
        self.assertFalse(at.exception, f"Page crashed: {at.exception}")
        self.assertEqual(at.title[0].value, "🔗 Actuals Integrator")

    def test_integrator_handles_invalid_tpm(self) -> None:
        """Uploading invalid TPM data should show an error."""
        at = AppTest.from_file(
            "actuals_integrator_page.py",
            default_timeout=30,
        )
        at.run()

        # Minimal valid summary
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shift A"  # type: ignore[union-attr]
        ws["C1"] = "11-05-2026"  # type: ignore[index]
        out = io.BytesIO()
        wb.save(out)

        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        at.file_uploader("integrator_summary").set_value(
            [("summary.xlsx", out.getvalue(), mime)],
        ).run()

        at.file_uploader("integrator_tpm").set_value(
            [("bad.xlsx", b"not excel", mime)],
        ).run()

        self.assertFalse(at.exception, f"App crashed: {at.exception}")
        self.assertGreater(len(at.error), 0)
        self.assertIn("Error integrating", at.error[0].value)

    def test_integrator_success(self) -> None:
        """Valid summary + valid TPM should show download button."""
        at = AppTest.from_file(
            "actuals_integrator_page.py",
            default_timeout=30,
        )
        at.run()

        # Build a minimal summary with a data row
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shift A"  # type: ignore[union-attr]
        ws["C1"] = "11-05-2026"  # type: ignore[index]
        ws.cell(row=8, column=2).value = "ACE COLT"  # type: ignore[union-attr]
        ws.cell(row=8, column=3).value = "J-001"  # type: ignore[union-attr]
        ws.cell(row=8, column=5).value = "766 0012 00 00 001"  # type: ignore[union-attr]
        ws.cell(row=8, column=7).value = "10"  # type: ignore[union-attr]
        s_out = io.BytesIO()
        wb.save(s_out)

        # Build a valid TPM file
        tpm_out = io.BytesIO()
        with pd.ExcelWriter(tpm_out, engine="openpyxl") as writer:
            blank = pd.DataFrame([[""] * 5] * 6)
            blank.to_excel(writer, index=False, header=False, startrow=0)  # type: ignore[reportUnknownMemberType]
            tpm_df = pd.DataFrame(
                {
                    "SHIFT": ["I"],
                    "DATE": ["20260511"],
                    "MACHINE NO": ["ACE COLT"],
                    "JOB ORDER NO.": ["J-001"],
                    "ACP QTY": [119],
                },
            )
            tpm_df.to_excel(writer, index=False, header=True, startrow=6)  # type: ignore[reportUnknownMemberType]

        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        at.file_uploader("integrator_summary").set_value(
            [("summary.xlsx", s_out.getvalue(), mime)],
        ).run()

        at.file_uploader("integrator_tpm").set_value(
            [("tpm.xlsx", tpm_out.getvalue(), mime)],
        ).run()

        self.assertFalse(at.exception, f"App crashed: {at.exception}")

        self.assertGreater(len(at.success), 0)
        has_download = any(
            getattr(e, "type", type(e).__name__) == "download_button" for e in at.main
        )
        self.assertTrue(has_download)
