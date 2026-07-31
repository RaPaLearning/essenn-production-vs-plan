"""Unit tests for the actuals_integrator module."""

import io
import unittest

import openpyxl
import pandas as pd

from actuals_integrator import process_actuals


class TestActualsIntegrator(unittest.TestCase):
    """Tests for process_actuals() and its helpers."""

    def _create_summary_bytes(self) -> bytes:
        """Build a minimal Operations Summary workbook."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Shift A"  # type: ignore[union-attr]

        ws["C1"] = "11-05-2026"  # type: ignore[index]

        headers = [
            "S.No.",
            "MACHINE",
            "JOB ORDER No",
            "TOTAL QTY",
            "PART NO",
            "PART NAME",
            "OPERATION",
            "PLAN QTY",
            "ACTUAL QTY",
        ]
        for c, h in enumerate(headers, 1):
            ws.cell(row=6, column=c).value = h  # type: ignore[union-attr]

        ws.cell(row=7, column=9).value = "OK QTY"  # type: ignore[union-attr]

        ws.cell(row=8, column=1).value = 1  # type: ignore[union-attr]
        ws.cell(row=8, column=2).value = "ACE COLT"  # type: ignore[union-attr]
        ws.cell(row=8, column=3).value = "J-001"  # type: ignore[union-attr]
        ws.cell(row=8, column=4).value = 1000  # type: ignore[union-attr]
        ws.cell(row=8, column=5).value = "766 0012 00 00 001"  # type: ignore[union-attr]
        ws.cell(row=8, column=6).value = "EMBASE DBM"  # type: ignore[union-attr]
        ws.cell(row=8, column=7).value = "10.0"  # type: ignore[union-attr]
        ws.cell(row=8, column=8).value = 500  # type: ignore[union-attr]

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    def _create_tpm_bytes(
        self,
        data: dict[str, list[object]],
    ) -> bytes:
        """Build a TPM file with 6 metadata rows + data."""
        out = io.BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            blank = pd.DataFrame([[""] * 5] * 6)
            blank.to_excel(  # type: ignore[reportUnknownMemberType]
                writer,
                index=False,
                header=False,
                startrow=0,
            )
            df = pd.DataFrame(data)
            df.to_excel(  # type: ignore[reportUnknownMemberType]
                writer,
                index=False,
                header=True,
                startrow=6,
            )
        return out.getvalue()

    def _get_default_tpm_data(self) -> dict[str, list[object]]:
        """Return a default TPM data dictionary for testing."""
        return {
            "SHIFT": ["Shift A"],
            "DATE": ["11/05/2026"],
            "MACHINE": ["ACE COLT"],
            "COMPONENT": ["766 0012 00 00 001"],
            "OPN NO.": ["10"],
            "ACTUAL": [119],
        }

    def _process_and_load(self, summary: bytes, tpm_files: list[bytes]) -> openpyxl.Workbook:
        """Helper to run the integrator and load the resulting workbook."""
        result = process_actuals(summary, tpm_files)
        return openpyxl.load_workbook(io.BytesIO(result))

    def test_matching_fills_ok_qty(self) -> None:
        """A matching TPM row should fill OK QTY."""
        summary = self._create_summary_bytes()
        tpm = self._create_tpm_bytes(self._get_default_tpm_data())
        wb = self._process_and_load(summary, [tpm])
        self.assertEqual(wb["Shift A"].cell(row=8, column=9).value, 119)

    def test_no_match_leaves_ok_qty_empty(self) -> None:
        """No matching TPM data leaves OK QTY untouched."""
        summary = self._create_summary_bytes()
        data = self._get_default_tpm_data()
        data.update(
            {
                "MACHINE": ["OTHER MACHINE"],
                "COMPONENT": ["DIFFERENT-PART"],
                "OPN NO.": ["20"],
                "ACTUAL": [999],
            }
        )
        tpm = self._create_tpm_bytes(data)
        wb = self._process_and_load(summary, [tpm])
        self.assertIsNone(wb["Shift A"].cell(row=8, column=9).value)

    def test_multiple_tpm_files_summed(self) -> None:
        """Actuals from multiple TPM files should be summed."""
        summary = self._create_summary_bytes()
        data1 = self._get_default_tpm_data()
        data1["ACTUAL"] = [100]
        tpm1 = self._create_tpm_bytes(data1)
        data2 = self._get_default_tpm_data()
        data2["ACTUAL"] = [50]
        tpm2 = self._create_tpm_bytes(data2)
        wb = self._process_and_load(summary, [tpm1, tpm2])
        self.assertEqual(wb["Shift A"].cell(row=8, column=9).value, 150)

    def test_invalid_tpm_raises(self) -> None:
        """If no TPM files can be parsed, ValueError is raised."""
        summary = self._create_summary_bytes()
        with self.assertRaises(Exception):
            process_actuals(summary, [b"not a valid excel file"])

    def test_shift_data_empty_skips(self) -> None:
        """TPM data for a different shift does not fill OK QTY."""
        summary = self._create_summary_bytes()
        data = self._get_default_tpm_data()
        data["SHIFT"] = ["Shift C"]
        data["ACTUAL"] = [100]
        tpm = self._create_tpm_bytes(data)
        wb = self._process_and_load(summary, [tpm])
        self.assertIsNone(wb["Shift A"].cell(row=8, column=9).value)

    def test_zero_actual_skips(self) -> None:
        """Zero actual total does not fill OK QTY."""
        summary = self._create_summary_bytes()
        data = self._get_default_tpm_data()
        data["ACTUAL"] = [0]
        tpm = self._create_tpm_bytes(data)
        wb = self._process_and_load(summary, [tpm])
        self.assertIsNone(wb["Shift A"].cell(row=8, column=9).value)

    def test_empty_part_row_skips(self) -> None:
        """Summary rows with no part number are skipped."""
        summary = self._create_summary_bytes()
        wb = openpyxl.load_workbook(io.BytesIO(summary))
        wb["Shift A"].cell(row=8, column=5).value = None
        out = io.BytesIO()
        wb.save(out)

        data = self._get_default_tpm_data()
        data["ACTUAL"] = [100]
        tpm = self._create_tpm_bytes(data)
        wb2 = self._process_and_load(out.getvalue(), [tpm])
        self.assertIsNone(wb2["Shift A"].cell(row=8, column=9).value)

    def test_missing_or_invalid_date_skips(self) -> None:
        """If date is missing or invalid in summary, the sheet is skipped."""
        data = self._get_default_tpm_data()
        data["ACTUAL"] = [100]
        tpm = self._create_tpm_bytes(data)

        # 1. Missing date
        wb = openpyxl.load_workbook(io.BytesIO(self._create_summary_bytes()))
        wb["Shift A"]["C1"] = None
        out = io.BytesIO()
        wb.save(out)
        wb1 = self._process_and_load(out.getvalue(), [tpm])
        self.assertIsNone(wb1["Shift A"].cell(row=8, column=9).value)

        # 2. Invalid date format
        wb = openpyxl.load_workbook(io.BytesIO(self._create_summary_bytes()))
        wb["Shift A"]["C1"] = "invalid-date"
        out2 = io.BytesIO()
        wb.save(out2)
        wb2 = self._process_and_load(out2.getvalue(), [tpm])
        self.assertIsNone(wb2["Shift A"].cell(row=8, column=9).value)

    def test_tpm_no_headers_skips(self) -> None:
        """If TPM lacks SHIFT/MACHINE headers, it is skipped."""
        out = io.BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df = pd.DataFrame({"NOT_SHIFT": ["A"], "NOT_MACHINE": ["B"]})
            df.to_excel(writer, index=False, header=True, startrow=0)  # type: ignore[reportUnknownMemberType]
        tpm_no_headers = out.getvalue()

        tpm_valid = self._create_tpm_bytes(self._get_default_tpm_data())

        summary = self._create_summary_bytes()
        # Mix the invalid one with a valid one so it doesn't just
        # raise ValueError for "no valid files"
        wb = self._process_and_load(summary, [tpm_no_headers, tpm_valid])
        self.assertEqual(wb["Shift A"].cell(row=8, column=9).value, 119)
