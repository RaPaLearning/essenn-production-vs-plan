"""Actuals Integrator – parse TPM IoT reports and fill OK QTY."""

import datetime
import io

import openpyxl
import openpyxl.worksheet.worksheet
import pandas as pd


def _find_header_row(raw: pd.DataFrame) -> int:
    """Find the row index containing SHIFT and MACHINE headers."""
    for idx, row in raw.iterrows():  # type: ignore[reportUnknownMemberType]
        vals = [str(v).upper().strip() for v in row if pd.notna(v)]
        if "SHIFT" in vals and "MACHINE" in vals:
            return int(idx)  # type: ignore[reportUnknownArgumentType]
    return -1


def _parse_single_tpm(tpm_bytes: bytes) -> pd.DataFrame | None:
    """Parse a single TPM file into a DataFrame, or None."""
    try:
        raw = pd.read_excel(  # type: ignore[reportUnknownMemberType]
            io.BytesIO(tpm_bytes),
            header=None,
        )
    except Exception:
        return None

    header_row = _find_header_row(raw)
    if header_row == -1:
        return None

    df: pd.DataFrame = pd.read_excel(  # type: ignore[reportUnknownMemberType]
        io.BytesIO(tpm_bytes),
        header=header_row,
    )
    df.columns = [str(c).upper().strip() for c in df.columns]

    return df.dropna(subset=["MACHINE"])


def _parse_all_tpm(tpm_bytes_list: list[bytes]) -> pd.DataFrame:
    """Parse all TPM files and return a combined DataFrame."""
    tpm_dfs: list[pd.DataFrame] = []
    for tpm_bytes in tpm_bytes_list:
        df = _parse_single_tpm(tpm_bytes)
        if df is not None:
            tpm_dfs.append(df)

    if not tpm_dfs:
        msg = "Could not parse valid TPM data from the uploaded files."
        raise ValueError(msg)

    tpm_data: pd.DataFrame = pd.concat(tpm_dfs, ignore_index=True)
    tpm_data["DATE"] = pd.to_datetime(  # type: ignore[reportUnknownMemberType]
        tpm_data["DATE"],
        dayfirst=True,
        format="mixed",
    ).dt.date
    return tpm_data


def _parse_sheet_date(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> datetime.date | None:
    """Parse the date from cell C1 of a summary sheet."""
    date_val = ws["C1"].value
    if not date_val:
        return None
    try:
        return datetime.datetime.strptime(
            str(date_val).strip(),
            "%d-%m-%Y",
        ).date()
    except ValueError:
        return None


def norm_str(s: str) -> str:
    """Normalize a string for fuzzy matching."""
    return str(s).replace("-", "").replace(" ", "").replace(".", "").upper()


def _match_component(
    matches: pd.DataFrame,
    part_str: str,
) -> pd.DataFrame:
    """Filter TPM rows by bidirectional substring match on component."""
    tpm_comps_norm = matches["COMPONENT"].astype(str).apply(norm_str)
    mask = tpm_comps_norm.apply(
        lambda tpm_c: (  # type: ignore[reportUnknownLambdaType]
            tpm_c != "NAN" and tpm_c != "" and (part_str in tpm_c or tpm_c in part_str)
        )
    )
    return matches[mask]


def _fill_row_actual(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    r_num: int,
    shift_data: pd.DataFrame,
) -> None:
    """Fill OK QTY for a single summary row if TPM data matches."""
    m_str = norm_str(str(ws.cell(row=r_num, column=2).value or ""))
    part_str = norm_str(str(ws.cell(row=r_num, column=5).value or ""))

    if not m_str or not part_str:
        return

    # Filter by Machine (Normalized exact match)
    tpm_machines_norm = shift_data["MACHINE"].astype(str).apply(norm_str)
    matches = _match_component(
        shift_data[tpm_machines_norm == m_str],
        part_str,
    )

    if not matches.empty:
        total_actual = int(pd.to_numeric(matches["ACTUAL"], errors="coerce").fillna(0).sum())
        if total_actual > 0:
            ws.cell(row=r_num, column=9).value = total_actual  # type: ignore[reportAttributeAccessIssue]


def _fill_sheet_actuals(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    tpm_data: pd.DataFrame,
) -> None:
    """Fill the OK QTY column for a single sheet."""
    target_date = _parse_sheet_date(ws)
    if target_date is None:
        return

    ws_title = str(ws.title).strip().upper()
    shift_data: pd.DataFrame = tpm_data[
        (tpm_data["DATE"] == target_date)
        & (tpm_data["SHIFT"].astype(str).str.strip().str.upper() == ws_title)
    ]
    if shift_data.empty:
        return

    for r_num in range(8, ws.max_row + 1):
        _fill_row_actual(ws, r_num, shift_data)


def process_actuals(
    summary_bytes: bytes,
    tpm_bytes_list: list[bytes],
) -> bytes:
    """Match TPM actuals to Operations Summary, fill OK QTY."""
    tpm_data = _parse_all_tpm(tpm_bytes_list)
    wb: openpyxl.Workbook = openpyxl.load_workbook(
        io.BytesIO(summary_bytes),
    )

    for sheet_name in wb.sheetnames:
        _fill_sheet_actuals(wb[sheet_name], tpm_data)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
