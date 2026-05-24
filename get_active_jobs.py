import datetime
from typing import Any
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from planquantity import compute_plan_qty

def _parse_date(row: "pd.Series[Any]", index: int) -> datetime.date | None:
    """Parse a date from the given column index of a row."""
    raw: object = row.iloc[index]
    if pd.isna(raw):  # type: ignore[arg-type]
        return None
    try:
        return pd.to_datetime(str(raw).strip(), dayfirst=True).date()  # type: ignore[union-attr]
    except Exception:
        return None


def _parse_datetime(row: "pd.Series[Any]", index: int) -> datetime.datetime | None:
    """Parse a full datetime from the given column index of a row."""
    raw: object = row.iloc[index]
    if pd.isna(raw):
        return None
    try:
        return pd.to_datetime(str(raw).strip(), dayfirst=True)
    except Exception:
        return None

def _process_row(row: "pd.Series[Any]", target: datetime.date) -> dict[str, Any] | None:
    """Return order details if the row's job is active on target date."""
    order: object = row.iloc[0]
    if pd.isna(order):  # type: ignore[arg-type]
        return None

    start: datetime.date | None = _parse_date(row, 8)
    end: datetime.date | None = _parse_date(row, 10)

    if start is None or end is None:
        return None

    if start <= target <= end:
        machine: object = row.iloc[7]
        machine_str: str = str(machine).strip() if not pd.isna(machine) else ""  # type: ignore[arg-type]

        # Extract template fields
        product: object = row.iloc[1]
        part_no: object = row.iloc[2]
        qty: object = row.iloc[3]
        op_name: object = row.iloc[6]

        qty_val = float(qty) if not pd.isna(qty) else 0.0
        part_no_str = str(part_no).strip() if not pd.isna(part_no) else ""
        product_str = str(product).strip() if not pd.isna(product) else ""
        op_name_str = str(op_name).strip() if not pd.isna(op_name) else ""

        start_dt = _parse_datetime(row, 8)
        end_dt = _parse_datetime(row, 10)
        plan_qty = compute_plan_qty(qty_val, start_dt, end_dt)

        return {
            "Machine": machine_str,
            "Job Order No": str(order).strip(),
            "Total Qty": int(qty_val),
            "Part No": part_no_str,
            "Part Name": product_str,
            "Operation": op_name_str,
            "Plan Qty": plan_qty
        }

    return None


def _process_sheet(df: pd.DataFrame, target: datetime.date) -> list[dict[str, Any]]:
    """Return all active job records from a single sheet."""
    if df.shape[0] < 6:
        return []

    results: list[dict[str, Any]] = []
    for _, row in df.iloc[5:].iterrows():
        record = _process_row(row, target)
        if record:
            results.append(record)
    return results

def _make_border(
    thin: Side,
    thick: Side,
    col: int,
    is_top_thick: bool,
    is_bottom_thick: bool,
) -> Border:
    """Create a cell border with thick edges on row/column boundaries."""
    return Border(
        left=thick if col == 1 else thin,
        right=thick if col == 13 else thin,
        top=thick if is_top_thick else thin,
        bottom=thick if is_bottom_thick else thin,
    )


def _write_header(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    label: str,
    target_date: datetime.date,
    thin: Side,
    thick: Side,
    meta_font: Font,
) -> None:
    """Write the DATE and SHIFT header rows."""
    formatted_date = target_date.strftime("%d-%m-%Y")
    bold_font = Font(name="Arial", size=10, bold=True)

    ws.merge_cells("A1:B1")
    ws["A1"] = "DATE:"
    ws["A1"].font = bold_font
    ws.merge_cells("C1:M1")
    ws["C1"] = formatted_date
    ws["C1"].font = meta_font
    ws["C1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:B2")
    ws["A2"] = "SHIFT:"
    ws["A2"].font = bold_font
    ws.merge_cells("C2:M2")
    ws["C2"] = f"{label} Shift"
    ws["C2"].font = meta_font
    ws["C2"].alignment = Alignment(horizontal="left", vertical="center")

    for r in [1, 2]:
        for c in range(1, 14):
            ws.cell(row=r, column=c).border = _make_border(
                thin, thick, c, r == 1, r == 2,
            )


def _style_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    r: int,
    font: Font,
    thin: Side,
    thick: Side,
    is_top_thick: bool,
    is_bottom_thick: bool,
    fill: PatternFill | None = None,
    align: Alignment | None = None,
) -> None:
    """Apply uniform styling to all 13 columns of a single row."""
    for c in range(1, 14):
        cell = ws.cell(row=r, column=c)
        cell.font = font
        cell.border = _make_border(
            thin, thick, c, is_top_thick, is_bottom_thick,
        )
        if fill is not None:
            cell.fill = fill
        if align is not None:
            cell.alignment = align


def _write_table_headers(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_font: Font,
    fill_grey: PatternFill,
    thin: Side,
    thick: Side,
) -> None:
    """Write the column headers in rows 6-7."""
    vertical_merges = [
        "A6:A7", "B6:B7", "C6:C7", "D6:D7", "E6:E7",
        "F6:F7", "G6:G7", "H6:H7", "K6:K7", "L6:L7", "M6:M7",
    ]
    for m in vertical_merges:
        ws.merge_cells(m)
    ws.merge_cells("I6:J6")

    headers = {
        "A6": "S.No.", "B6": "MACHINE", "C6": "JOB ORDER No",
        "D6": "TOTAL QTY", "E6": "PART NO", "F6": "PART NAME",
        "G6": "OPERATION", "H6": "PLAN QTY", "I6": "ACTUAL QTY",
        "I7": "OK QTY", "J7": "Rej QTY", "K6": "REMARKS",
        "L6": "Operator Name", "M6": "Sign",
    }
    for cell_ref, value in headers.items():
        ws[cell_ref] = value

    center = Alignment(
        horizontal="center", vertical="center", wrap_text=True,
    )
    for r in [6, 7]:
        ws.row_dimensions[r].height = 25
        _style_row(
            ws, r, header_font, thin, thick,
            r == 6, r == 7, fill_grey, center,
        )


def _write_single_data_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    r_num: int,
    idx: int,
    row: dict[str, Any],
    data_font: Font,
    thin: Side,
    thick: Side,
    is_last: bool,
) -> None:
    """Write one job data row and style its cells."""
    center_cols = {1, 4, 8, 9, 10}
    data_keys = [
        "Machine", "Job Order No", "Total Qty",
        "Part No", "Part Name", "Operation", "Plan Qty",
    ]
    ws.row_dimensions[r_num].height = 24
    ws.cell(row=r_num, column=1, value=idx + 1)
    for col_i, key in enumerate(data_keys, start=2):
        ws.cell(row=r_num, column=col_i, value=row[key])
    for c in range(9, 14):
        ws.cell(row=r_num, column=c, value="")
    for c in range(1, 14):
        cell = ws.cell(row=r_num, column=c)
        cell.font = data_font
        cell.border = _make_border(thin, thick, c, False, is_last)
        h = "center" if c in center_cols else "left"
        cell.alignment = Alignment(horizontal=h, vertical="center")


def _write_empty_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    data_font: Font,
    thin: Side,
    thick: Side,
) -> None:
    """Write a placeholder row when no jobs are active."""
    ws.row_dimensions[start_row].height = 24
    ws.cell(row=start_row, column=1, value="-")
    ws.cell(
        row=start_row, column=2,
        value="No active jobs scheduled for this shift",
    )
    ws.merge_cells(
        start_row=start_row, start_column=2,
        end_row=start_row, end_column=13,
    )
    _style_row(
        ws, start_row, data_font, thin, thick,
        False, True,
    )


def _write_data_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    rows: list[dict[str, Any]],
    data_font: Font,
    thin: Side,
    thick: Side,
) -> int:
    """Write job data rows starting at row 8. Returns last grid row."""
    start_row = 8
    for idx, row in enumerate(rows):
        _write_single_data_row(
            ws, start_row + idx, idx, row,
            data_font, thin, thick,
            is_last=(idx == len(rows) - 1),
        )

    if rows:
        return start_row + len(rows) - 1

    _write_empty_row(ws, start_row, data_font, thin, thick)
    return start_row


def _write_footer(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    last_grid_row: int,
) -> None:
    """Write signature footer and set column widths."""
    foot_row = last_grid_row + 3
    sign_font = Font(name="Arial", size=10, bold=True)
    ws.cell(
        row=foot_row, column=2, value="Sign Of Incharge :",
    ).font = sign_font
    ws.cell(
        row=foot_row, column=9, value="Sign Of Manager :",
    ).font = sign_font

    col_widths = {
        1: 6, 2: 18, 3: 16, 4: 12, 5: 14, 6: 28, 7: 18,
        8: 12, 9: 10, 10: 10, 11: 16, 12: 18, 13: 10,
    }
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


def _fill_shift_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    label: str,
    target_date: datetime.date,
    rows: list[dict[str, Any]],
) -> None:
    """Style and fill the given sheet with the shift template layout."""
    ws.views.sheetView[0].showGridLines = True

    header_font = Font(name="Arial", size=10, bold=True)
    meta_font = Font(name="Arial", size=9)
    data_font = Font(name="Arial", size=10)
    thin = Side(style='thin', color='000000')
    thick = Side(style='medium', color='000000')
    fill_grey = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid",
    )

    _write_header(ws, label, target_date, thin, thick, meta_font)
    _write_table_headers(ws, header_font, fill_grey, thin, thick)
    last_row = _write_data_rows(ws, rows, data_font, thin, thick)
    _write_footer(ws, last_row)


def export_active_jobs(input_path: str, date_str: str, output_path: str) -> None:
    """Export active jobs and pre-fill shift-wise template sheets to the output Excel file."""
    target: datetime.date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
    sheets: dict[str, pd.DataFrame] = pd.read_excel(  # type: ignore[reportUnknownMemberType]
        input_path, sheet_name=None, header=None
    )

    rows: list[dict[str, Any]] = []
    for df in sheets.values():
        rows.extend(_process_sheet(df, target))

    # Save workbook with exactly the three shift sheets
    wb = openpyxl.Workbook()

    # Configure Shift I on the default first sheet
    ws_shift1 = wb.active
    ws_shift1.title = "Shift I"
    _fill_shift_sheet(ws_shift1, "1st", target, rows)

    # Create Shift II and III sheets
    ws_shift2 = wb.create_sheet(title="Shift II")
    _fill_shift_sheet(ws_shift2, "2nd", target, rows)

    ws_shift3 = wb.create_sheet(title="Shift III")
    _fill_shift_sheet(ws_shift3, "3rd", target, rows)

    wb.save(output_path)
